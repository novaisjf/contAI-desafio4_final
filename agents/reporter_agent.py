import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
from .context import Contexto

class ReporterAgent:
    """
    Agente que formata e gera o arquivo de saída final.
    """

    def execute(self, base_calculada: pd.DataFrame, bases: dict, ctx: Contexto, out_xlsx: str) -> float:
        """
        Recebe a base final calculada e a exporta para uma planilha Excel formatada.
        """
        logging.info("Agente Relator: Iniciando geração do relatório final.")
        if base_calculada.empty:
            logging.warning("Agente Relator: Base calculada está vazia. Nenhum relatório será gerado.")
            return 0.0

        df = base_calculada.copy()

        final_cols = [
            "Matricula", "Admissão", "Sindicato do Colaborador", "Competência", "Dias", 
            "VALOR DIÁRIO VR", "TOTAL", "Custo empresa", "Desconto profissional", "OBS GERAL"
        ]
        
        final_df = pd.DataFrame({
            "Matricula": df["MATRICULA"].astype("Int64"),
            "Admissão": df["ADMISSAO"],
            "Sindicato do Colaborador": df["SINDICATO"],
            "Competência": ctx.competencia.strftime('%m/%Y'),
            "Dias": df["DIAS_CALCULADOS"].astype(int),
            "VALOR DIÁRIO VR": df["VALOR_UNITARIO"].astype(float),
            "TOTAL": df["VR_TOTAL"].astype(float),
            "Custo empresa": df["EMPRESA_80"].astype(float),
            "Desconto profissional": df["COLABORADOR_20"].astype(float),
            "OBS GERAL": df["OBS GERAL"],
        })
        final_df = final_df[final_cols]

        # --- Lógica para a aba de Validações ---
        valor_total_vr = final_df["TOTAL"].sum()
        des = bases.get("DESLIGADOS", pd.DataFrame())
        des_ok_ate15 = 0
        des_pos15 = 0
        if not des.empty and "DATA DEMISSÃO" in des.columns and "OK" in des.columns:
            des_ok_ate15 = ((des["DATA DEMISSÃO"].dt.day <= 15) & (des["OK"])).sum()
            des_pos15 = (des["DATA DEMISSÃO"].dt.day >= 16).sum()

        sv = bases.get("SIND_VALOR", pd.DataFrame())
        sind_resumo = ", ".join(f"{r.ESTADO}: {r.VALOR:.2f}" for _, r in sv.iterrows())

        valid_lines = [
            ("VALOR TOTAL VR", f"R$ {valor_total_vr:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
            ("Colaboradores Processados", len(final_df)),
            ("---", "---"),
            ("Afastados / Licenças", len(bases.get("AFASTAMENTOS", pd.DataFrame()))),
            ("DESLIGADOS GERAL", len(des)),
            ("Admitidos mês", len(bases.get("ADMISSAO", pd.DataFrame()))),
            ("Férias (total dias)", int(bases.get("FERIAS", pd.DataFrame()).get("DIAS DE FÉRIAS", 0).sum())),
            ("ESTAGIARIO", len(bases.get("ESTAGIO", pd.DataFrame()))),
            ("APRENDIZ", len(bases.get("APRENDIZ", pd.DataFrame()))),
            ("SINDICATOS x VALOR", sind_resumo),
            ("DESLIGADOS ATÉ O DIA 15 DO MÊS", int(des_ok_ate15)),
            ("DESLIGADOS DO DIA 16 EM DIANTE", int(des_pos15)),
            ("EXTERIOR", len(bases.get("EXTERIOR", pd.DataFrame()))),
            ("ATIVOS (base original)", len(bases.get("ATIVOS", pd.DataFrame()))),
        ]
        valid_df = pd.DataFrame(valid_lines, columns=["Validações","Check"])

        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
            sheet_name = f"VR MENSAL {ctx.competencia.strftime('%m.%Y')}"
            final_df.to_excel(w, sheet_name=sheet_name, index=False)
            valid_df.to_excel(w, sheet_name="Validações", index=False)

        # --- Ajustes de formatação com openpyxl ---
        wb = openpyxl.load_workbook(out_xlsx)
        ws = wb[sheet_name]

    # Cabeçalhos permanecem na primeira linha, sem linha de totalização

        # 2. Formatação dos headers (linha 2)
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")


        wb.save(out_xlsx)

        logging.info(f"Agente Relator: Relatório final salvo em '{out_xlsx}'. Valor total: {valor_total_vr}")
        return valor_total_vr